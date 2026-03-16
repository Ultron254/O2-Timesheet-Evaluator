from __future__ import annotations

import difflib
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "Department",
    "Employee",
    "Month",
    "Date",
    "Client",
    "Product",
    "Task",
    "Notes",
    "Hours",
    "Reviewer's Comments",
]

REQUIRED_MIN_COLUMNS = {"Employee", "Date", "Hours"}

COLUMN_ALIASES = {
    "Department": [
        "department",
        "deparrtment",
        "dept",
        "team",
        "unit",
        "division",
    ],
    "Employee": [
        "employee",
        "employee name",
        "name",
        "staff",
        "consultant",
        "resource",
    ],
    "Month": ["month", "period", "timesheet month", "month-year", "year month"],
    "Date": ["date", "work date", "entry date", "day", "timesheet date"],
    "Client": ["client", "account", "customer", "brand"],
    "Product": ["product", "service", "service line", "category", "workstream"],
    "Task": ["task", "activity", "work done", "description", "assignment"],
    "Notes": ["notes", "remarks", "comment", "details", "narration"],
    "Hours": ["hours", "hrs", "time", "hours worked", "duration", "effort"],
    "Reviewer's Comments": [
        "reviewer's comments",
        "reviewer comments",
        "review comment",
        "review",
        "approval comment",
        "manager comments",
    ],
}

PREFERRED_SHEETS = ["Original", "Timesheet", "Data"]
DATE_FORMATS = [
    "%d-%b-%Y",
    "%d/%m/%Y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%m/%d/%Y",
    "%d %b %Y",
    "%b %d %Y",
]


def _norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _clean_text(series: pd.Series) -> pd.Series:
    return (
        series.fillna("")
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )


def _score_header(column_name: str, canonical: str) -> float:
    col = _norm(column_name)
    aliases = [_norm(alias) for alias in COLUMN_ALIASES[canonical]]
    best = 0.0
    for alias in aliases:
        if not alias:
            continue
        if col == alias:
            return 1.0
        if alias in col or col in alias:
            best = max(best, 0.92)
        else:
            best = max(best, difflib.SequenceMatcher(None, col, alias).ratio())
    return best


def _looks_like_date(series: pd.Series) -> float:
    parsed = pd.to_datetime(_clean_text(series), errors="coerce", dayfirst=False)
    ratio = float(parsed.notna().mean())
    if ratio < 0.5:
        parsed_day_first = pd.to_datetime(_clean_text(series), errors="coerce", dayfirst=True)
        ratio = max(ratio, float(parsed_day_first.notna().mean()))
    return ratio


def _looks_like_hours(series: pd.Series) -> float:
    numeric = pd.to_numeric(series, errors="coerce")
    numeric_ratio = float(numeric.notna().mean())
    if numeric_ratio <= 0:
        return 0.0
    bounded_ratio = float(((numeric >= 0) & (numeric <= 24 * 2)).mean())
    return 0.6 * numeric_ratio + 0.4 * bounded_ratio


def _looks_like_month(series: pd.Series) -> float:
    text = _clean_text(series).str.lower()
    month_pattern = (
        r"jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|"
        r"january|february|march|april|june|july|august|september|"
        r"october|november|december"
    )
    return float(text.str.contains(month_pattern, regex=True, na=False).mean())


def _looks_like_reviewer_comments(series: pd.Series) -> float:
    text = _clean_text(series).str.lower()
    review_tokens = r"\bok\b|check|hours look large|kinda ok|exess hours|approve|review"
    return float(text.str.contains(review_tokens, regex=True, na=False).mean())


def _detect_content_columns(df: pd.DataFrame, mapped: Dict[str, str]) -> Dict[str, str]:
    used = set(mapped.values())
    remaining = [c for c in df.columns if c not in used]

    def pick_best(score_fn, threshold: float = 0.55) -> Optional[str]:
        best_col: Optional[str] = None
        best_score = threshold
        for col in remaining:
            score = score_fn(df[col])
            if score > best_score:
                best_score = score
                best_col = col
        if best_col:
            remaining.remove(best_col)
            used.add(best_col)
        return best_col

    if "Hours" not in mapped:
        candidate = pick_best(_looks_like_hours, threshold=0.60)
        if candidate:
            mapped["Hours"] = candidate

    if "Date" not in mapped:
        candidate = pick_best(_looks_like_date, threshold=0.55)
        if candidate:
            mapped["Date"] = candidate

    if "Month" not in mapped:
        candidate = pick_best(_looks_like_month, threshold=0.45)
        if candidate:
            mapped["Month"] = candidate

    if "Reviewer's Comments" not in mapped:
        candidate = pick_best(_looks_like_reviewer_comments, threshold=0.25)
        if candidate:
            mapped["Reviewer's Comments"] = candidate

    if "Employee" not in mapped:
        best_col: Optional[str] = None
        best_score = 0.45
        for col in remaining:
            series = _clean_text(df[col])
            non_empty_ratio = float((series != "").mean())
            unique_ratio = float(series.nunique(dropna=True) / max(len(series), 1))
            mean_len = float(series.str.len().mean() or 0.0)
            score = (0.50 * non_empty_ratio) + (0.35 * min(unique_ratio, 1.0)) + (0.15 * min(mean_len / 20.0, 1.0))
            if score > best_score:
                best_col = col
                best_score = score
        if best_col:
            mapped["Employee"] = best_col
            remaining.remove(best_col)

    return mapped


def detect_and_map_columns(df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    used: set[str] = set()

    for canonical in EXPECTED_COLUMNS:
        best_col = None
        best_score = 0.72
        for col in df.columns:
            if col in used:
                continue
            score = _score_header(col, canonical)
            if score > best_score:
                best_col = col
                best_score = score
        if best_col is not None:
            mapping[canonical] = best_col
            used.add(best_col)

    mapping = _detect_content_columns(df, mapping)
    missing_required = [col for col in REQUIRED_MIN_COLUMNS if col not in mapping]
    if missing_required:
        raise ValueError(
            "Could not detect required columns: "
            f"{', '.join(missing_required)}. At minimum Employee, Date, and Hours are required."
        )
    return mapping


def _choose_sheet(file_path: Path) -> Optional[str]:
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        xl = pd.ExcelFile(file_path, engine="openpyxl")
        sheet_names = list(xl.sheet_names)
    else:
        xl = pd.ExcelFile(file_path)
        sheet_names = list(xl.sheet_names)
    if not sheet_names:
        return None
    normalized = {_norm(name): name for name in sheet_names}
    for preferred in PREFERRED_SHEETS:
        match = normalized.get(_norm(preferred))
        if match:
            return match
    return sheet_names[0]


def _uniquify_headers(headers: List[str]) -> List[str]:
    counts: Dict[str, int] = {}
    out: List[str] = []
    for raw in headers:
        name = str(raw or "").strip()
        if not name:
            name = "Unnamed"
        base = name
        if base not in counts:
            counts[base] = 0
            out.append(base)
            continue
        counts[base] += 1
        out.append(f"{base}_{counts[base]}")
    return out


def _read_excel_with_openpyxl(path: Path, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header_row: Optional[Tuple[object, ...]] = None
    for row in rows_iter:
        if row is None:
            continue
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header_row = row
            break
    if header_row is None:
        return pd.DataFrame()

    headers = _uniquify_headers([str(cell or "").strip() for cell in header_row])
    width = len(headers)

    records: List[List[object]] = []
    source_indices: List[int] = []
    source_row = ws.min_row + 1
    for row in rows_iter:
        source_row += 1
        if row is None:
            continue
        values = list(row[:width])
        if len(values) < width:
            values.extend([None] * (width - len(values)))
        if all(val is None or str(val).strip() == "" for val in values):
            continue
        records.append(values)
        source_indices.append(source_row)

    df = pd.DataFrame(records, columns=headers)
    # Keep source row index consistent with legacy behavior.
    df.insert(0, "source_row_index", source_indices)
    return df


def _parse_dates(date_series: pd.Series, month_series: pd.Series) -> pd.Series:
    date_text = _clean_text(date_series)
    parsed = pd.to_datetime(date_text, errors="coerce", dayfirst=False)
    missing = parsed.isna()
    if missing.any():
        parsed_dayfirst = pd.to_datetime(date_text[missing], errors="coerce", dayfirst=True)
        parsed.loc[missing] = parsed_dayfirst

    for fmt in DATE_FORMATS:
        missing = parsed.isna()
        if not missing.any():
            break
        parsed_fmt = pd.to_datetime(date_text[missing], format=fmt, errors="coerce")
        parsed.loc[missing] = parsed_fmt

    missing = parsed.isna()
    if missing.any():
        month_text = _clean_text(month_series[missing])
        month_parsed = pd.to_datetime(month_text, errors="coerce")
        parsed.loc[missing] = month_parsed

    return parsed


def _remove_total_rows(df: pd.DataFrame) -> pd.DataFrame:
    date_txt = _clean_text(df["Date"]).str.lower()
    month_txt = _clean_text(df["Month"]).str.lower()
    client_txt = _clean_text(df["Client"]).str.lower()
    product_txt = _clean_text(df["Product"]).str.lower()
    task_txt = _clean_text(df["Task"]).str.lower()
    notes_txt = _clean_text(df["Notes"]).str.lower()

    is_total_literal = (
        date_txt.eq("total")
        | month_txt.eq("total")
        | client_txt.eq("total")
    )
    all_total_like = (
        client_txt.str.contains("total", na=False)
        & product_txt.str.contains("total", na=False)
        & task_txt.str.contains("total", na=False)
        & notes_txt.str.contains("total", na=False)
    )
    return df.loc[~(is_total_literal | all_total_like)].copy()


def load_and_clean_timesheet(file_path: str | Path) -> pd.DataFrame:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Timesheet file not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw_df = pd.read_csv(path)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        sheet_name = _choose_sheet(path)
        if sheet_name is None:
            raise ValueError("No sheet found in uploaded workbook.")
        try:
            raw_df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        except Exception:
            raw_df = _read_excel_with_openpyxl(path, sheet_name=sheet_name)
    elif suffix == ".xls":
        sheet_name = _choose_sheet(path)
        if sheet_name is None:
            raise ValueError("No sheet found in uploaded workbook.")
        raw_df = pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
    else:
        raise ValueError(f"Unsupported file format: {suffix}")

    if raw_df.empty:
        raise ValueError("Uploaded file is empty.")

    raw_df = raw_df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if "source_row_index" not in raw_df.columns:
        raw_df = raw_df.reset_index(drop=False).rename(columns={"index": "source_row_index"})

    mapped = detect_and_map_columns(raw_df)

    cleaned = pd.DataFrame()
    for canonical in EXPECTED_COLUMNS:
        source = mapped.get(canonical)
        if source is None:
            cleaned[canonical] = ""
        else:
            cleaned[canonical] = raw_df[source]

    cleaned["source_row_index"] = raw_df["source_row_index"]

    for col in ["Department", "Employee", "Month", "Date", "Client", "Product", "Task", "Notes", "Reviewer's Comments"]:
        cleaned[col] = _clean_text(cleaned[col])

    cleaned["Hours"] = pd.to_numeric(cleaned["Hours"], errors="coerce")
    cleaned["Hours"] = cleaned["Hours"].fillna(0.0)

    cleaned = _remove_total_rows(cleaned)
    cleaned = cleaned[cleaned["Employee"] != ""].copy()

    cleaned["date_parsed"] = _parse_dates(cleaned["Date"], cleaned["Month"])
    cleaned = cleaned[cleaned["date_parsed"].notna()].copy()
    if cleaned.empty:
        raise ValueError("No valid timesheet rows remained after cleaning and date parsing.")

    cleaned["day_of_week"] = cleaned["date_parsed"].dt.dayofweek.astype(int)
    cleaned["is_weekend"] = cleaned["day_of_week"] >= 5
    cleaned["week_number"] = cleaned["date_parsed"].dt.isocalendar().week.astype(int)
    cleaned["year"] = cleaned["date_parsed"].dt.year.astype(int)
    cleaned["year_month"] = cleaned["date_parsed"].dt.to_period("M").astype(str)

    client_lower = cleaned["Client"].str.lower()
    task_lower = cleaned["Task"].str.lower()

    cleaned["is_leave"] = task_lower.str.contains("leave", na=False)
    cleaned["is_holiday"] = task_lower.str.contains("holiday", na=False)
    cleaned["is_internal"] = client_lower.str.contains("internal", na=False) | cleaned["is_leave"] | cleaned["is_holiday"]
    cleaned["work_type"] = np.where(cleaned["is_internal"], "Internal/Overhead", "Billable")

    cleaned["Month"] = cleaned["Month"].where(cleaned["Month"] != "", cleaned["date_parsed"].dt.strftime("%B %Y"))

    return cleaned.reset_index(drop=True)
